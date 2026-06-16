import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import io
import re

# ------------------------------------------------------------------
# CONFIGURACAO DA PAGINA
# ------------------------------------------------------------------
st.set_page_config(
    page_title="Migrador de Produtos - ITECWEB",
    page_icon="??",
    layout="wide"
)

# ------------------------------------------------------------------
# CONSTANTES
# ------------------------------------------------------------------
COLUNAS_ERP = [
    {"nome": "Codigo",                    "obrigatorio": True,  "tipo": "numero", "max_chars": 10},
    {"nome": "Descricao",                 "obrigatorio": True,  "tipo": "texto",  "max_chars": 200, "chars_proibidos": ["'", '"', "=", "&"]},
    {"nome": "Referencia",                "obrigatorio": False, "tipo": "texto",  "max_chars": 50},
    {"nome": "Marca",                     "obrigatorio": False, "tipo": "texto",  "max_chars": 50},
    {"nome": "Secao",                     "obrigatorio": False, "tipo": "texto",  "max_chars": 100},
    {"nome": "Grupo",                     "obrigatorio": False, "tipo": "texto",  "max_chars": 200},
    {"nome": "NCM",                       "obrigatorio": True,  "tipo": "texto",  "digitos_exatos": 8},
    {"nome": "CEST",                      "obrigatorio": False, "tipo": "texto",  "digitos_exatos": 7},
    {"nome": "Tipo Item",                 "obrigatorio": True,  "tipo": "texto",  "digitos_exatos": 2, "opcoes": ["00","01","02","03","04","05","06","07","08","09","99"]},
    {"nome": "Codigo Barra",              "obrigatorio": False, "tipo": "texto",  "max_chars": 30},
    {"nome": "Preco Venda",               "obrigatorio": True,  "tipo": "numero", "max_decimais": 2},
    {"nome": "Custo",                     "obrigatorio": True,  "tipo": "numero", "max_decimais": 2},
    {"nome": "Margem Lucro",              "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Estoque",                   "obrigatorio": True,  "tipo": "numero", "max_decimais": 2},
    {"nome": "Estoque Minimo",            "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Estoque Maximo",            "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Unidade Medida",            "obrigatorio": True,  "tipo": "texto",  "max_chars": 50},
    {"nome": "Origem",                    "obrigatorio": True,  "tipo": "numero", "digitos_exatos": 1, "opcoes": ["0","1","2","3","4","5","6","7","8"]},
    {"nome": "CFOP Saida",                "obrigatorio": True,  "tipo": "numero", "digitos_exatos": 4},
    {"nome": "CST ICMS",                  "obrigatorio": True,  "tipo": "texto",  "max_chars": 3},
    {"nome": "ALIQUOTA ICMS",             "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "REDUCAO BASE ICMS",         "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CST ICMS NFCE",             "obrigatorio": True,  "tipo": "texto",  "max_chars": 3},
    {"nome": "ALIQUOTA ICMS NFCE",        "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "REDUCAO BASE ICMS NFCE",    "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CFOP NFCE",                 "obrigatorio": True,  "tipo": "numero", "digitos_exatos": 4},
    {"nome": "CST IPI",                   "obrigatorio": False, "tipo": "texto",  "digitos_exatos": 2},
    {"nome": "ALIQUOTA IPI",              "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CST PIS",                   "obrigatorio": True,  "tipo": "texto",  "digitos_exatos": 2},
    {"nome": "ALIQUOTA PIS",              "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CST COFINS",                "obrigatorio": True,  "tipo": "texto",  "digitos_exatos": 2},
    {"nome": "ALIQUOTA COFINS",           "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "ALIQUOTA DIFERIMENTO",      "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "ALIQUOTA SIMPLES NACIONAL", "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CST ICMS SAT",              "obrigatorio": False, "tipo": "texto",  "max_chars": 3},
    {"nome": "OBSERVACAO",                "obrigatorio": False, "tipo": "texto",  "max_chars": 500},
    {"nome": "NATUREZA DE RECEITA",       "obrigatorio": False, "tipo": "texto",  "max_chars": 10},
    {"nome": "CODIGO ANP",                "obrigatorio": False, "tipo": "texto",  "max_chars": 12},
    {"nome": "LOCAL ESTOQUE",             "obrigatorio": False, "tipo": "texto",  "max_chars": 100},
    {"nome": "PESO BRUTO",                "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "PESO LIQUIDO",              "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "CST ICMS ST",               "obrigatorio": False, "tipo": "texto",  "max_chars": 3},
    {"nome": "CFOP ST",                   "obrigatorio": False, "tipo": "numero", "digitos_exatos": 4},
    {"nome": "CFOP ST Interestadual",     "obrigatorio": False, "tipo": "numero", "digitos_exatos": 4},
    {"nome": "Modalidade BC ICMS ST",     "obrigatorio": False, "tipo": "texto",  "max_chars": 1, "opcoes": ["0","1","2","3"]},
    {"nome": "MVA Pauta",                 "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Reducao BC ICMS ST",        "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Aliquota ICMS Interna",     "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
    {"nome": "Reducao BC ICMS Interna",   "obrigatorio": False, "tipo": "numero", "max_decimais": 2},
]

NOMES_COLUNAS_ERP = [c["nome"] for c in COLUNAS_ERP]

OPCOES_TIPO_ITEM = {
    "00 - Mercadoria para Revenda": "00",
    "01 - Materia-Prima": "01",
    "02 - Embalagem": "02",
    "03 - Produto em Processo": "03",
    "04 - Produto Acabado": "04",
    "05 - Subproduto": "05",
    "06 - Produto Intermediario": "06",
    "07 - Material de Uso e Consumo": "07",
    "08 - Ativo Imobilizado": "08",
    "09 - Servicos": "09",
    "99 - Outros Insumos": "99",
}

OPCOES_ORIGEM = {
    "0 - Nacional": "0",
    "1 - Estrangeira importacao direta": "1",
    "2 - Estrangeira adquirida mercado interno": "2",
    "3 - Nacional conteudo importacao acima 40%": "3",
    "4 - Nacional producao conforme processos basicos": "4",
    "5 - Nacional conteudo importacao abaixo 40%": "5",
    "6 - Estrangeira importacao direta sem similar": "6",
    "7 - Estrangeira adquirida interno sem similar": "7",
    "8 - Nacional conteudo importacao acima 70%": "8",
}

CAMPOS_COM_OPCOES = {
    "Tipo Item": OPCOES_TIPO_ITEM,
    "Origem":    OPCOES_ORIGEM,
}

DEFAULTS_FIXO = {
    "CST PIS":    "99",
    "CST COFINS": "99",
}

CAMPOS_MODO_FIXO_PADRAO = {
    "CST PIS", "CST COFINS", "Tipo Item", "Origem",
    "CST ICMS", "CST ICMS NFCE", "CFOP Saida", "CFOP NFCE"
}

VALORES_PADRAO_VAZIO = {
    "NCM":            "00000000",
    "Unidade Medida": "unidade",
    "Preco Venda":    "0.00",
    "Estoque":        "0",
    "Custo":          "0.00",
}

CAMPOS_UNIFICADOS = {
    "CFOP (Saida e NFCe)":     ["CFOP Saida", "CFOP NFCE"],
    "CST ICMS (NFe e NFCe)":   ["CST ICMS",   "CST ICMS NFCE"],
}

CAMPOS_OCULTOS_POR_UNIFICADO = {"CFOP Saida", "CFOP NFCE", "CST ICMS", "CST ICMS NFCE"}

# ------------------------------------------------------------------
# VALIDACAO
# ------------------------------------------------------------------
def validar_celula(valor, config):
    erros = []
    nome_campo = config["nome"]

    vazio = (
        valor is None or
        str(valor).strip() == "" or
        str(valor).strip().lower() in ("nan", "none", "nat")
    )

    if vazio:
        if nome_campo in VALORES_PADRAO_VAZIO:
            valor = VALORES_PADRAO_VAZIO[nome_campo]
        elif nome_campo == "Descricao":
            erros.append("Campo obrigatorio vazio")
            return erros, ""
        else:
            if config.get("obrigatorio"):
                erros.append("Campo obrigatorio vazio")
            return erros, ""

    val_str = str(valor).strip()

    if config["tipo"] == "numero":
        try:
            f = float(val_str.replace(",", "."))
            if "digitos_exatos" in config:
                val_str = str(int(f))
            elif config.get("max_decimais") == 0:
                val_str = str(int(f))
            else:
                val_str = str(round(f, config.get("max_decimais", 2)))
        except Exception:
            erros.append("Valor invalido para numero: " + val_str)
            return erros, val_str
    else:
        if "digitos_exatos" in config:
            digitos   = config["digitos_exatos"]
            val_limpo = re.sub(r"\D", "", val_str)
            if len(val_limpo) != digitos:
                erros.append(
                    "Deve ter exatamente " + str(digitos) +
                    " digitos. Encontrado: " + val_str
                )
            else:
                val_str = val_limpo

        if "opcoes" in config:
            if val_str not in config["opcoes"]:
                erros.append(
                    "Valor invalido: " + val_str +
                    ". Opcoes: " + str(config["opcoes"])
                )

        if "max_chars" in config:
            if len(val_str) > config["max_chars"]:
                val_str = val_str[:config["max_chars"]]
                erros.append("Truncado para " + str(config["max_chars"]) + " caracteres")

        if "chars_proibidos" in config:
            for ch in config["chars_proibidos"]:
                if ch in val_str:
                    val_str = val_str.replace(ch, "")

    return erros, val_str


def salvar_excel_formatado(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan1"

    amarelo      = PatternFill("solid", fgColor="FFD700")
    azul         = PatternFill("solid", fgColor="AED6F1")
    obrigatorios = {c["nome"] for c in COLUNAS_ERP if c["obrigatorio"]}

    for col_idx, nome_col in enumerate(NOMES_COLUNAS_ERP, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.fill      = amarelo if nome_col in obrigatorios else azul
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(len(nome_col) + 4, 16)

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(row_data, start=1):
            col_cfg = COLUNAS_ERP[col_idx - 1]
            cell    = ws.cell(row=row_idx, column=col_idx)
            val_str = str(valor).strip() if valor else ""

            if col_cfg["tipo"] == "numero" and val_str and val_str.lower() != "nan":
                try:
                    num = float(val_str.replace(",", "."))
                    if col_cfg.get("digitos_exatos") or col_cfg.get("max_decimais") == 0:
                        cell.value         = int(num)
                        cell.number_format = "0"
                    else:
                        cell.value         = round(num, 2)
                        cell.number_format = "0.00"
                except Exception:
                    cell.value = val_str
            else:
                cell.value = val_str if val_str and val_str.lower() != "nan" else None

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ------------------------------------------------------------------
# INICIALIZA SESSION STATE
# ------------------------------------------------------------------
def init_state():
    defaults = {
        "etapa":               1,
        "arquivos":            [],
        "arquivo_idx":         0,
        "mapeamentos":         {},
        "mapeamentos_unif":    {},
        "codigo_modo":         "sequencial",
        "codigo_seq_inicio":   "1",
        "codigo_coluna":       "(nao mapear)",
        "log_linhas":          [],
        "arquivo_gerado":      None,
        "nome_arquivo_gerado": "",
        "processado":          False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ------------------------------------------------------------------
# ESTILOS - SEM ACENTOS, BOTOES TODOS AZUL
# ------------------------------------------------------------------
st.markdown("""
<style>
    .titulo-principal {
        font-size: 26px; font-weight: bold;
        color: #1a5276; margin-bottom: 4px;
    }
    .subtitulo {
        font-size: 13px; color: #555; margin-bottom: 20px;
    }
    .etapa-badge {
        background: #2980b9; color: white;
        border-radius: 20px; padding: 4px 16px;
        font-size: 13px; font-weight: bold;
        display: inline-block; margin-bottom: 12px;
    }
    .regra-box {
        background: #eafaf1; border-left: 4px solid #27ae60;
        padding: 10px 16px; border-radius: 4px;
        font-size: 13px; color: #1e8449; margin-bottom: 16px;
    }
    .log-box {
        background: #1e1e1e; color: #d4d4d4;
        font-family: monospace; font-size: 12px;
        padding: 12px; border-radius: 6px;
        height: 320px; overflow-y: auto; white-space: pre-wrap;
    }
    .unificado-box {
        background: #eaf4fb; border-left: 4px solid #2980b9;
        padding: 8px 14px; border-radius: 4px;
        font-size: 12px; color: #1a5276; margin-top: 4px;
    }
    /* Todos os botoes padrao em azul */
    div[data-testid="stButton"] button {
        background-color: #2e86c1 !important;
        color: white !important;
        border: none !important;
        font-weight: bold !important;
    }
    div[data-testid="stButton"] button:hover {
        background-color: #1a5276 !important;
        color: white !important;
    }
    /* Botao primary tambem azul */
    div[data-testid="stButton"] button[kind="primary"] {
        background-color: #1a5276 !important;
        color: white !important;
        border: none !important;
        font-weight: bold !important;
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        background-color: #154360 !important;
    }
    /* Botao download azul */
    div[data-testid="stDownloadButton"] button {
        background-color: #1a5276 !important;
        color: white !important;
        font-size: 15px !important;
        font-weight: bold !important;
        border: none !important;
        width: 100% !important;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background-color: #154360 !important;
    }
</style>
""", unsafe_allow_html=True)

# ------------------------------------------------------------------
# CABECALHO
# ------------------------------------------------------------------
st.markdown('<div class="titulo-principal">&#128230; Migrador de Produtos - ITECWEB</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitulo">Converta planilhas de origem para o layout padrao do ERP</div>', unsafe_allow_html=True)

# ------------------------------------------------------------------
# NAVEGACAO
# ------------------------------------------------------------------
cols_nav = st.columns(3)
etapas_nav = ["1 - Selecionar Arquivos", "2 - Mapear Campos", "3 - Gerar Planilha"]
for i, (col, nome) in enumerate(zip(cols_nav, etapas_nav), start=1):
    with col:
        ativo = st.session_state.etapa == i
        cor   = "#1a5276" if ativo else "#aab7b8"
        peso  = "bold" if ativo else "normal"
        st.markdown(
            f'<div style="background:{cor};color:white;padding:8px;'
            f'border-radius:6px;text-align:center;font-size:13px;'
            f'font-weight:{peso}">{nome}</div>',
            unsafe_allow_html=True
        )
st.markdown("---")

# ------------------------------------------------------------------
# HELPER VALIDACAO OBRIGATORIOS
# ------------------------------------------------------------------
def campos_obrigatorios_faltando(arquivos, mapeamentos, mapeamentos_unif, codigo_modo, codigo_coluna):
    faltando = []
    multiplos = len(arquivos) > 1

    if codigo_modo == "coluna":
        if multiplos:
            faltando.append("Codigo: modo COLUNA indisponivel para multiplas planilhas")
        else:
            for arq in arquivos:
                col = codigo_coluna
                if not col or col == "(nao mapear)" or col not in arq["df"].columns:
                    faltando.append(f"Codigo: coluna '{col}' invalida no arquivo '{arq['nome']}'")

    for arq in arquivos:
        nome_arq = arq["nome"]
        df       = arq["df"]
        mapa     = mapeamentos.get(nome_arq, {})
        mapa_u   = mapeamentos_unif.get(nome_arq, {})

        for nome_unif, filhos in CAMPOS_UNIFICADOS.items():
            cfg_u = mapa_u.get(nome_unif, {})
            modo  = cfg_u.get("modo", "fixo")
            if modo == "fixo":
                val = cfg_u.get("fixo", "").strip()
                if not val:
                    faltando.append(f"{nome_unif}: arquivo '{nome_arq}' valor fixo vazio")
            else:
                col = cfg_u.get("coluna", "(nao mapear)")
                if not col or col == "(nao mapear)" or col not in df.columns:
                    faltando.append(f"{nome_unif}: arquivo '{nome_arq}' nenhuma coluna valida selecionada")

        for col_cfg in COLUNAS_ERP:
            campo = col_cfg["nome"]
            if not col_cfg.get("obrigatorio"):
                continue
            if campo == "Codigo":
                continue
            if campo in CAMPOS_OCULTOS_POR_UNIFICADO:
                continue
            if campo in VALORES_PADRAO_VAZIO:
                continue

            cfg_campo = mapa.get(campo, {})
            modo      = cfg_campo.get("modo", "coluna")

            if modo == "fixo":
                val = cfg_campo.get("fixo", "").strip()
                if not val:
                    faltando.append(f"{campo}: arquivo '{nome_arq}' valor fixo vazio")
            else:
                col = cfg_campo.get("coluna", "(nao mapear)")
                if not col or col == "(nao mapear)" or col not in df.columns:
                    faltando.append(f"{campo}: arquivo '{nome_arq}' nenhuma coluna valida selecionada")

    return faltando


# ==================================================================
# ETAPA 1
# ==================================================================
if st.session_state.etapa == 1:
    st.markdown('<div class="etapa-badge">Etapa 1 - Selecionar Arquivos</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="regra-box">
    <b>Regras automaticas aplicadas na geracao:</b><br>
    &bull; Descricao vazia: linha descartada da planilha final<br>
    &bull; Descricao com ' " = &: caracteres removidos automaticamente<br>
    &bull; NCM vazio: preenchido com 00000000<br>
    &bull; Unidade Medida vazia: preenchida com unidade<br>
    &bull; Preco Venda / Custo vazios: preenchidos com 0,00<br>
    &bull; Estoque vazio: preenchido com 0<br>
    &bull; Multiplas planilhas: Codigo sempre sequencial continuo
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Selecione uma ou mais planilhas (.xlsx ou .csv)",
        type=["xlsx", "csv"],
        accept_multiple_files=True,
        key="uploader_etapa1"
    )

    if uploaded:
        arquivos_carregados = []
        erros_carga         = []

        for uf in uploaded:
            try:
                if uf.name.lower().endswith(".csv"):
                    df = pd.read_csv(uf, dtype=str)
                else:
                    df = pd.read_excel(uf, dtype=str)

                df.columns = [str(c).strip() for c in df.columns]

                mapeamento = {}
                for col in COLUNAS_ERP:
                    nome = col["nome"]
                    if nome == "Codigo" or nome in CAMPOS_OCULTOS_POR_UNIFICADO:
                        continue
                    mapeamento[nome] = {
                        "modo":   "fixo" if nome in CAMPOS_MODO_FIXO_PADRAO else "coluna",
                        "fixo":   DEFAULTS_FIXO.get(nome, ""),
                        "coluna": "(nao mapear)",
                    }

                mapeamento_unif = {}
                for nome_unif in CAMPOS_UNIFICADOS:
                    mapeamento_unif[nome_unif] = {
                        "modo":   "fixo",
                        "fixo":   "",
                        "coluna": "(nao mapear)",
                    }

                arquivos_carregados.append({"nome": uf.name, "df": df})
                st.session_state.mapeamentos[uf.name]      = mapeamento
                st.session_state.mapeamentos_unif[uf.name] = mapeamento_unif

            except Exception as e:
                erros_carga.append(f"{uf.name} -> {e}")

        if erros_carga:
            for e in erros_carga:
                st.error(e)

        if arquivos_carregados:
            st.session_state.arquivos    = arquivos_carregados
            st.session_state.arquivo_idx = 0
            st.success(f"{len(arquivos_carregados)} arquivo(s) carregado(s) com sucesso.")
            for arq in arquivos_carregados:
                st.write(f"OK **{arq['nome']}** - {len(arq['df'])} linhas, {len(arq['df'].columns)} colunas")

    if st.session_state.arquivos:
        st.markdown("---")
        if st.button("Proximo: Mapear Campos >>", type="primary", use_container_width=True):
            st.session_state.etapa = 2
            st.rerun()


# ==================================================================
# ETAPA 2
# ==================================================================
elif st.session_state.etapa == 2:
    st.markdown('<div class="etapa-badge">Etapa 2 - Mapear e Configurar Campos</div>', unsafe_allow_html=True)

    arquivos  = st.session_state.arquivos
    multiplos = len(arquivos) > 1
    nomes     = [a["nome"] for a in arquivos]

    nome_sel = st.selectbox("Configurando arquivo:", nomes, index=st.session_state.arquivo_idx)
    st.session_state.arquivo_idx = nomes.index(nome_sel)
    arq_atual = arquivos[st.session_state.arquivo_idx]
    df_atual  = arq_atual["df"]
    mapa      = st.session_state.mapeamentos[nome_sel]
    mapa_u    = st.session_state.mapeamentos_unif[nome_sel]

    colunas_origem = ["(nao mapear)"] + list(df_atual.columns)

    st.caption(
        f"{len(df_atual)} linhas | {len(df_atual.columns)} colunas | "
        f"Colunas: {', '.join(df_atual.columns.tolist()[:8])}"
        f"{'...' if len(df_atual.columns) > 8 else ''}"
    )

    filtro = st.radio("Mostrar campos:", ["Todos", "Somente obrigatorios"], horizontal=True)
    st.markdown("---")

    # CODIGO
    st.markdown("### Configuracao do Codigo")
    c1, c2 = st.columns([1, 2])
    with c1:
        if multiplos:
            st.info("Multiplas planilhas: Codigo sera Sequencial automatico.")
            st.session_state.codigo_modo = "sequencial"
            modo_cod = "sequencial"
        else:
            modo_cod = st.radio(
                "Modo do Codigo:",
                ["sequencial", "coluna", "fixo"],
                index=["sequencial", "coluna", "fixo"].index(st.session_state.codigo_modo),
                format_func=lambda x: {"sequencial": "Sequencial", "coluna": "Coluna", "fixo": "Fixo"}[x]
            )
            st.session_state.codigo_modo = modo_cod
    with c2:
        if modo_cod == "sequencial":
            val = st.text_input("Inicio do sequencial:", value=st.session_state.codigo_seq_inicio)
            st.session_state.codigo_seq_inicio = val
            st.caption("Numera automaticamente todos os produtos")
        elif modo_cod == "coluna":
            col_cod = st.selectbox(
                "Coluna do Codigo:", colunas_origem,
                index=colunas_origem.index(st.session_state.codigo_coluna)
                if st.session_state.codigo_coluna in colunas_origem else 0
            )
            st.session_state.codigo_coluna = col_cod
        else:
            val = st.text_input("Valor fixo do Codigo:", value=st.session_state.codigo_seq_inicio)
            st.session_state.codigo_seq_inicio = val

    st.markdown("---")

    # CAMPOS UNIFICADOS
    st.markdown("### Campos Unificados (preenchem dois campos na saida)")
    for nome_unif, filhos in CAMPOS_UNIFICADOS.items():
        label_exp = f"OBRIGATORIO | {nome_unif} - preenche: {filhos[0]} e {filhos[1]}"
        with st.expander(label_exp, expanded=True):
            c1, c2 = st.columns([1, 2])
            cfg_u  = mapa_u.get(nome_unif, {"modo": "fixo", "fixo": "", "coluna": "(nao mapear)"})

            with c1:
                novo_modo = st.radio(
                    "Modo:",
                    ["coluna", "fixo"],
                    index=0 if cfg_u.get("modo") == "coluna" else 1,
                    key=f"modo_unif_{nome_sel}_{nome_unif}",
                    format_func=lambda x: "Coluna da planilha" if x == "coluna" else "Valor Fixo"
                )
                cfg_u["modo"] = novo_modo

            with c2:
                if novo_modo == "coluna":
                    col_atual = cfg_u.get("coluna", "(nao mapear)")
                    idx_col   = colunas_origem.index(col_atual) if col_atual in colunas_origem else 0
                    col_sel   = st.selectbox(
                        "Coluna de origem:", colunas_origem, index=idx_col,
                        key=f"col_unif_{nome_sel}_{nome_unif}"
                    )
                    cfg_u["coluna"] = col_sel
                    if col_sel != "(nao mapear)" and col_sel in df_atual.columns:
                        exemplos = df_atual[col_sel].dropna().astype(str).head(3).tolist()
                        st.caption("Exemplos: " + " / ".join(exemplos))
                else:
                    val_fixo = st.text_input(
                        "Valor fixo (mesmo para ambos os campos):",
                        value=cfg_u.get("fixo", ""),
                        key=f"fixo_unif_{nome_sel}_{nome_unif}"
                    )
                    cfg_u["fixo"] = val_fixo

                st.markdown(
                    f'<div class="unificado-box">Este valor sera copiado para: '
                    f'<b>{filhos[0]}</b> e <b>{filhos[1]}</b></div>',
                    unsafe_allow_html=True
                )

            mapa_u[nome_unif] = cfg_u

    st.session_state.mapeamentos_unif[nome_sel] = mapa_u

    st.markdown("---")
    st.markdown("### Demais Campos")
    st.markdown("Estrela = Obrigatorio | Quadrado = Opcional")

    for col_cfg in COLUNAS_ERP:
        nome  = col_cfg["nome"]
        if nome == "Codigo":
            continue
        if nome in CAMPOS_OCULTOS_POR_UNIFICADO:
            continue
        obrig = col_cfg.get("obrigatorio", False)
        if filtro == "Somente obrigatorios" and not obrig:
            continue

        tem_padrao = nome in VALORES_PADRAO_VAZIO
        prefixo    = "[*] " if obrig else "[ ] "
        label      = f"{prefixo}{nome}"
        if obrig:
            label += " (OBRIGATORIO)"
        if tem_padrao:
            label += f" | padrao: {VALORES_PADRAO_VAZIO[nome]}"

        with st.expander(label, expanded=False):
            c1, c2    = st.columns([1, 2])
            cfg_campo = mapa.get(nome, {"modo": "coluna", "fixo": "", "coluna": "(nao mapear)"})

            with c1:
                modo_atual = cfg_campo.get("modo", "coluna")
                novo_modo  = st.radio(
                    "Modo:",
                    ["coluna", "fixo"],
                    index=0 if modo_atual == "coluna" else 1,
                    key=f"modo_{nome_sel}_{nome}",
                    format_func=lambda x: "Coluna da planilha" if x == "coluna" else "Valor Fixo"
                )
                cfg_campo["modo"] = novo_modo

            with c2:
                if novo_modo == "coluna":
                    col_atual = cfg_campo.get("coluna", "(nao mapear)")
                    idx_col   = colunas_origem.index(col_atual) if col_atual in colunas_origem else 0
                    col_sel   = st.selectbox(
                        "Coluna de origem:", colunas_origem, index=idx_col,
                        key=f"col_{nome_sel}_{nome}"
                    )
                    cfg_campo["coluna"] = col_sel
                    if col_sel != "(nao mapear)" and col_sel in df_atual.columns:
                        exemplos = df_atual[col_sel].dropna().astype(str).head(3).tolist()
                        st.caption("Exemplos: " + " / ".join(exemplos))
                else:
                    if nome in CAMPOS_COM_OPCOES:
                        opcoes_dict  = CAMPOS_COM_OPCOES[nome]
                        opcoes_lista = list(opcoes_dict.keys())
                        val_atual    = cfg_campo.get("fixo", "")
                        idx_op       = 0
                        for i, k in enumerate(opcoes_lista):
                            if opcoes_dict[k] == val_atual or k == val_atual:
                                idx_op = i
                                break
                        sel_op = st.selectbox(
                            "Valor fixo:", opcoes_lista, index=idx_op,
                            key=f"fixo_{nome_sel}_{nome}"
                        )
                        cfg_campo["fixo"] = opcoes_dict[sel_op]
                    else:
                        val_fixo = st.text_input(
                            "Valor fixo (para todos os produtos):",
                            value=cfg_campo.get("fixo", ""),
                            key=f"fixo_{nome_sel}_{nome}"
                        )
                        cfg_campo["fixo"] = val_fixo

                regras = []
                if col_cfg.get("digitos_exatos"):
                    regras.append(f"{col_cfg['digitos_exatos']} digitos exatos")
                if col_cfg.get("max_chars"):
                    regras.append(f"max {col_cfg['max_chars']} caracteres")
                if col_cfg.get("max_decimais") is not None:
                    regras.append(f"{col_cfg['max_decimais']} casas decimais")
                if tem_padrao:
                    regras.append(f"se vazio: {VALORES_PADRAO_VAZIO[nome]}")
                if regras:
                    st.caption("Regras: " + " | ".join(regras))

            mapa[nome] = cfg_campo

    st.session_state.mapeamentos[nome_sel] = mapa

    st.markdown("---")
    col_v, col_p = st.columns(2)
    with col_v:
        if st.button("<< Voltar: Selecionar Arquivos", use_container_width=True):
            st.session_state.etapa = 1
            st.rerun()
    with col_p:
        if st.button("Proximo: Gerar Planilha >>", type="primary", use_container_width=True):
            st.session_state.etapa = 3
            st.rerun()


# ==================================================================
# ETAPA 3
# ==================================================================
elif st.session_state.etapa == 3:
    st.markdown('<div class="etapa-badge">Etapa 3 - Gerar Planilha</div>', unsafe_allow_html=True)

    arquivos = st.session_state.arquivos
    st.info(
        f"{len(arquivos)} arquivo(s) carregado(s) | "
        f"{sum(len(a['df']) for a in arquivos)} linhas no total"
    )

    faltando = campos_obrigatorios_faltando(
        arquivos,
        st.session_state.mapeamentos,
        st.session_state.mapeamentos_unif,
        st.session_state.codigo_modo,
        st.session_state.codigo_coluna,
    )

    if faltando:
        itens = "".join(f"<br>&bull; {f}" for f in faltando)
        st.markdown(
            f'<div style="background:#c0392b;color:white;padding:12px 16px;'
            f'border-radius:6px;font-size:13px;font-weight:bold;margin-bottom:12px;">'
            f'ITENS OBRIGATORIOS FALTANDO:{itens}</div>',
            unsafe_allow_html=True
        )

    st.markdown("---")
    col_v, col_g = st.columns(2)
    with col_v:
        if st.button("<< Voltar: Mapear Campos", use_container_width=True):
            st.session_state.etapa      = 2
            st.session_state.processado = False
            st.rerun()
    with col_g:
        gerar = st.button(
            "GERAR PLANILHA >>",
            type="primary",
            use_container_width=True,
            disabled=bool(faltando)
        )

    if faltando:
        st.caption("Configure os campos obrigatorios antes de gerar.")

    st.markdown("---")

    if gerar:
        st.session_state.processado     = False
        st.session_state.arquivo_gerado = None
        st.session_state.log_linhas     = []

        modo_cod = st.session_state.codigo_modo
        try:
            seq_inicio = int(st.session_state.codigo_seq_inicio.strip())
        except Exception:
            seq_inicio = 1

        linhas_resultado   = []
        erros_totais       = 0
        avisos_totais      = 0
        erros_por_campo    = {}
        linhas_descartadas = 0
        contador_seq       = seq_inicio
        log                = []

        def add_log(msg):
            log.append(msg)

        add_log("INICIO DO PROCESSAMENTO - ITECWEB")
        add_log(
            f"Codigo: SEQUENCIAL inicio {seq_inicio}"
            if modo_cod == "sequencial"
            else f"Codigo: {modo_cod.upper()}"
        )
        add_log("-" * 60)

        barra    = st.progress(0, text="Processando...")
        total_ln = sum(len(a["df"]) for a in arquivos)
        proc     = 0

        for arq in arquivos:
            df       = arq["df"]
            nome_arq = arq["nome"]
            mapa     = st.session_state.mapeamentos.get(nome_arq, {})
            mapa_u   = st.session_state.mapeamentos_unif.get(nome_arq, {})

            add_log(f">> Arquivo: {nome_arq} ({len(df)} linhas) <<")

            for idx, row in df.iterrows():
                linha_erp  = {}
                desc_vazia = False

                valores_unificados = {}
                for nome_unif, filhos in CAMPOS_UNIFICADOS.items():
                    cfg_u = mapa_u.get(nome_unif, {})
                    if cfg_u.get("modo") == "coluna":
                        col = cfg_u.get("coluna", "(nao mapear)")
                        val = str(row.get(col, "")) if col in df.columns else ""
                    else:
                        val = cfg_u.get("fixo", "")
                    if str(val).lower() in ("nan", "none", "nat"):
                        val = ""
                    for filho in filhos:
                        valores_unificados[filho] = val

                for col_cfg in COLUNAS_ERP:
                    campo = col_cfg["nome"]

                    if campo == "Codigo":
                        if modo_cod == "sequencial":
                            valor = str(contador_seq)
                        elif modo_cod == "coluna":
                            col   = st.session_state.codigo_coluna
                            valor = str(row.get(col, "")) if col in df.columns else ""
                        else:
                            valor = st.session_state.codigo_seq_inicio.strip()

                    elif campo in CAMPOS_OCULTOS_POR_UNIFICADO:
                        valor = valores_unificados.get(campo, "")

                    else:
                        cfg_campo = mapa.get(campo, {})
                        if cfg_campo.get("modo") == "fixo":
                            valor = cfg_campo.get("fixo", "")
                        else:
                            col   = cfg_campo.get("coluna", "(nao mapear)")
                            valor = str(row.get(col, "")) if (col and col != "(nao mapear)" and col in df.columns) else ""

                    if str(valor).lower() in ("nan", "none", "nat"):
                        valor = ""

                    erros, valor_tratado = validar_celula(valor, col_cfg)

                    for erro in erros:
                        nivel = "ERRO" if col_cfg["obrigatorio"] else "AVISO"
                        add_log(f"[{nivel}] {nome_arq} | Linha {idx + 2} | {campo}: {erro}")
                        if nivel == "ERRO":
                            erros_totais += 1
                            erros_por_campo[campo] = erros_por_campo.get(campo, 0) + 1
                        else:
                            avisos_totais += 1

                    if campo == "Descricao":
                        val_final = valor_tratado if valor_tratado != "" else valor
                        if not str(val_final).strip():
                            desc_vazia = True

                    linha_erp[campo] = valor_tratado if valor_tratado != "" else valor

                if desc_vazia:
                    linhas_descartadas += 1
                    add_log(f"[DESCARTADO] {nome_arq} | Linha {idx + 2} | Descricao vazia")
                    proc += 1
                    barra.progress(min(proc / total_ln, 1.0), text=f"Processando... {proc}/{total_ln}")
                    continue

                if modo_cod == "sequencial":
                    contador_seq += 1

                linhas_resultado.append(linha_erp)
                proc += 1
                barra.progress(min(proc / total_ln, 1.0), text=f"Processando... {proc}/{total_ln}")

        barra.progress(1.0, text="Concluido!")

        df_final = pd.DataFrame(linhas_resultado, columns=NOMES_COLUNAS_ERP)
        buffer   = salvar_excel_formatado(df_final)

        add_log("-" * 60)
        add_log(f"CONCLUIDO - {len(df_final)} produtos incluidos")
        add_log(f"Linhas descartadas (desc. vazias): {linhas_descartadas}")
        add_log(f"Erros: {erros_totais}")
        add_log(f"Avisos: {avisos_totais}")

        if erros_por_campo:
            add_log("")
            add_log("RESUMO DE ERROS POR CAMPO:")
            for campo, qtd in sorted(erros_por_campo.items(), key=lambda x: -x[1]):
                add_log(f"  - {campo}: {qtd} erro(s)")

        st.session_state.log_linhas          = log
        st.session_state.arquivo_gerado      = buffer.getvalue()
        st.session_state.nome_arquivo_gerado = "migracao_produtos_itecweb.xlsx"
        st.session_state.processado          = True
        st.rerun()

    if st.session_state.processado and st.session_state.arquivo_gerado:
        log      = st.session_state.log_linhas
        erros_t  = sum(1 for l in log if l.startswith("[ERRO]"))
        avisos_t = sum(1 for l in log if l.startswith("[AVISO]"))
        desc_t   = sum(1 for l in log if l.startswith("[DESCARTADO]"))

        try:
            total_prod = int([l for l in log if "CONCLUIDO" in l][0].split(" - ")[1].split(" ")[0])
        except Exception:
            total_prod = 0

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Produtos incluidos", total_prod)
        c2.metric("Descartados",        desc_t)
        c3.metric("Erros",              erros_t)
        c4.metric("Avisos",             avisos_t)

        st.download_button(
            label="Baixar planilha gerada (.xlsx)",
            data=st.session_state.arquivo_gerado,
            file_name=st.session_state.nome_arquivo_gerado,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.markdown("#### Log de processamento")
        log_texto = "\n".join(log)
        st.markdown(f'<div class="log-box">{log_texto}</div>', unsafe_allow_html=True)

        if erros_t == 0:
            st.success("Planilha gerada sem erros! Pronta para importar.")
        else:
            st.warning(f"Planilha gerada com {erros_t} erro(s). Verifique o log antes de importar.")